import os

import pandas as pd
import yaml
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

yamlPath = 'R:/iuye/config/application.yml'

with open(yamlPath, 'rb') as f:
    config_data = yaml.safe_load(f)

iuye_data_config = config_data['iuye_data']


def is_file_exist(file_name):
    return os.path.isfile(file_name)


def get_data_input_path(relative_path=None):
    return get_data_path('input_path', relative_path)


def get_data_output_path(relative_path=None):
    return get_data_path('output_path', relative_path)


def get_data_archive_path(relative_path=None):
    return get_data_path('archive_path', relative_path)


def get_data_path(file_type, relative_path=None):
    file_path = iuye_data_config[file_type]
    if relative_path is None:
        return file_path
    else:
        data_path = '%s%s' % (file_path, relative_path)
        if not os.path.exists(data_path):
            os.makedirs(data_path)
        return data_path


def get_config_path(relative_path=None):
    config_path = config_data['config_path']
    if relative_path is not None:
        config_path = '%s%s' % (config_path, relative_path)
    return config_path


def reset_col(filename):
    wb = load_workbook(filename)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        df = pd.read_excel(filename, sheet).fillna('-')
        if not df.empty:
            df.loc[len(df)] = list(df.columns)  # 把标题行附加到最后一行
            for col in df.columns:
                index = list(df.columns).index(col)  # 列序号
                letter = get_column_letter(index + 1)  # 列字母
                collen = df[col].apply(lambda x: len(str(x).encode())).max()  # 获取这一列长度的最大值 当然也可以用min获取最小值 mean获取平均值
                ws.column_dimensions[letter].width = collen * 1.1 + 1

    wb.save(filename)


if __name__ == '__main__':
    print(get_data_output_path())
    print(get_config_path())
